import openpyxl
def max_row(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_row
def max_column(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_column
def read_data(file,sheetname,row_number,col_number):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row_number,col_number).value
def write_data(file,sheetname,row_number,col_number,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row_number,col_number).value = data
    workbook.save(file)
    
